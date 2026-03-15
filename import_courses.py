#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
课表导入工具 - 支持多种格式的矩阵式课表

使用方法：
    python import_courses.py 学生课表.xlsx

支持的格式：
    格式1（教务系统格式，带编号）：
        课程名[编码] 数字
        周次 节次 教师[编码,职称] 地点
    
    格式2（简化格式，推荐）：
        课程名[编号]
        周次 教师地点
    
    示例：
        高等数学[001]
1-16周 示例老师A 示例教室A
        
        大学物理[002]
1-16周(单) 示例老师B 示例教室B
"""

import os
import sys
import json
import re
import pandas as pd
from typing import List, Dict, Any
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# 节次映射：行索引 -> (开始节, 结束节)
SECTION_MAPPING = {
    2: (1, 2), 3: (1, 2),
    4: (3, 4), 5: (3, 4),
    6: (5, 6), 7: (5, 6),
    8: (7, 8), 9: (7, 8),
    10: (9, 10), 11: (9, 10),
}

# 星期映射
WEEKDAY_MAPPING = {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7}

WEEKDAY_LABELS = {
    '周一': 1,
    '星期一': 1,
    '周二': 2,
    '星期二': 2,
    '周三': 3,
    '星期三': 3,
    '周四': 4,
    '星期四': 4,
    '周五': 5,
    '星期五': 5,
    '周六': 6,
    '星期六': 6,
    '周日': 7,
    '星期日': 7,
    '星期天': 7,
}

BIG_SECTION_INDEX = {
    '一': 1,
    '二': 2,
    '三': 3,
    '四': 4,
    '五': 5,
    '六': 6,
    '七': 7,
    '八': 8,
    '九': 9,
    '十': 10,
}


def normalize_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace('\r\n', '\n').replace('\r', '\n').strip()


def build_dataframe_from_excel(file_path: str) -> pd.DataFrame:
    ext = os.path.splitext(file_path)[1].lower()
    if load_workbook is None or ext == '.xls':
        return pd.read_excel(file_path, header=None)

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        matrix = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]

        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = ws.cell(min_row, min_col).value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    matrix[r - 1][c - 1] = top_left_value

        return pd.DataFrame(matrix)
    except Exception:
        return pd.read_excel(file_path, header=None)


def detect_weekday_columns(df: pd.DataFrame) -> Dict[int, int]:
    detected = {}
    scan_rows = min(len(df), 5)

    for row_idx in range(scan_rows):
        for col_idx in range(len(df.columns)):
            text = normalize_cell_text(df.iloc[row_idx, col_idx])
            if not text:
                continue
            for label, weekday in WEEKDAY_LABELS.items():
                if label in text:
                    detected[col_idx] = weekday
                    break

    if len(detected) >= 5:
        return detected

    return WEEKDAY_MAPPING.copy()


def parse_section_from_label(text: str) -> Any:
    if not text:
        return None

    range_match = re.search(r'第?\s*(\d+)\s*[-~到]\s*(\d+)\s*节', text)
    if range_match:
        return int(range_match.group(1)), int(range_match.group(2))

    big_match = re.search(r'第\s*([一二三四五六七八九十\d]+)\s*大节', text)
    if big_match:
        token = big_match.group(1)
        index = int(token) if token.isdigit() else BIG_SECTION_INDEX.get(token)
        if index:
            start = (index - 1) * 2 + 1
            end = start + 1
            return start, end

    return None


def build_row_section_mapping(df: pd.DataFrame) -> Dict[int, Any]:
    mapping = {}

    for row_idx in range(len(df)):
        first_col_text = normalize_cell_text(df.iloc[row_idx, 0]) if len(df.columns) > 0 else ""
        section = parse_section_from_label(first_col_text)
        if section:
            mapping[row_idx] = section

    if mapping:
        return mapping

    return SECTION_MAPPING.copy()


def parse_arrangements_detailed(text: str) -> List[Dict[str, Any]]:
    """
    解析课程安排信息（教务系统格式，带教师编号）
    格式：16周 5-6节 教师[编码,职称] 地点
    """
    results = []
    pattern = r'(\d+)(?:-(\d+))?周(?:\((单|双)\))?\s+(\d+)-(\d+)节\s+([^\[\s]+)\[(\d+),([^\]]+)\]\s+(\S+)'
    
    for match in re.finditer(pattern, text):
        start_week = int(match.group(1))
        end_week = int(match.group(2)) if match.group(2) else start_week
        week_type_str = match.group(3)
        start_section = int(match.group(4))
        end_section = int(match.group(5))
        teacher_name = match.group(6).strip()
        location = match.group(9).strip()
        
        if '班' in location:
            location = re.sub(r'[\u4e00-\u9fa5]+\d*班.*', '', location).strip()
        
        week_type = 'all'
        if week_type_str == '单':
            week_type = 'odd'
        elif week_type_str == '双':
            week_type = 'even'
        
        results.append({
            'start_week': start_week,
            'end_week': end_week,
            'week_type': week_type,
            'start_section': start_section,
            'end_section': end_section,
            'teacher': teacher_name,
            'location': location
        })
    
    return results


def parse_arrangements_simple(text: str, default_start: int, default_end: int) -> List[Dict[str, Any]]:
    """
    解析课程安排信息（简化格式）
    格式：1-16周 教师姓名 地点
    或：1-16周(单) 教师姓名 地点
    """
    results = []
    
    # 匹配周次（支持单双周）
    week_pattern = r'(\d+)(?:-(\d+))?周(?:\((单|双)\))?'
    
    for match in re.finditer(week_pattern, text):
        start_week = int(match.group(1))
        end_week = int(match.group(2)) if match.group(2) else start_week
        week_type_str = match.group(3)
        
        week_type = 'all'
        if week_type_str == '单':
            week_type = 'odd'
        elif week_type_str == '双':
            week_type = 'even'
        
        # 提取周次后面的内容（教师和地点）
        after_week = text[match.end():].strip()
        
        # 尝试匹配教师和地点
        # 格式：教师姓名 地点
        teacher = ""
        location = ""
        
        # 常见地点模式
        loc_patterns = [
            r'(一教\d+|二教\d+|三教\d+|科研楼\d+|外语楼\d+|机房\d+|操场|体育馆|\S+楼\d+|\S+场)',
        ]
        
        for loc_pat in loc_patterns:
            loc_match = re.search(loc_pat, after_week)
            if loc_match:
                location = loc_match.group(1)
                # 教师是地点之前的内容
                teacher_part = after_week[:loc_match.start()].strip()
                # 清理教师名（去掉可能的班级信息）
                teacher = re.sub(r'[\u4e00-\u9fa5]+\d*班', '', teacher_part).strip()
                break
        
        if not location:
            # 没有匹配到地点，把周次后面的内容当作教师
            teacher = after_week.split()[0] if after_week.split() else ""
        
        results.append({
            'start_week': start_week,
            'end_week': end_week,
            'week_type': week_type,
            'start_section': default_start,
            'end_section': default_end,
            'teacher': teacher,
            'location': location
        })
    
    return results


def parse_course_cell(cell_content: str, weekday: int, default_start: int, default_end: int) -> List[Dict[str, Any]]:
    """解析单元格中的课程"""
    if not cell_content or pd.isna(cell_content):
        return []
    
    content = str(cell_content).strip()
    if not content:
        return []
    
    courses = []
    
    # 按空行分割不同的课程块
    blocks = re.split(r'\n\s*\n', content)
    
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        
        lines = [line.strip() for line in block.split('\n') if line.strip()]
        first_line = lines[0] if lines else ""

        course_match = re.match(r'^([^\[\n]+?)(?:\[[^\]\n]*\])?(?:\s+\d+)?$', first_line)
        
        if course_match:
            course_name = course_match.group(1).strip()
            rest = '\n'.join(lines[1:]).strip()
            
            # 先尝试详细格式（带节次）
            arrangements = parse_arrangements_detailed(rest)
            
            # 如果详细格式没匹配到，尝试简化格式
            if not arrangements:
                arrangements = parse_arrangements_simple(rest, default_start, default_end)
            
            for arr in arrangements:
                courses.append({
                    'course_name': course_name,
                    'teacher': arr['teacher'],
                    'location': arr['location'],
                    'weekday': weekday,
                    'start_section': arr['start_section'],
                    'end_section': arr['end_section'],
                    'week_type': arr['week_type'],
                    'start_week': arr['start_week'],
                    'end_week': arr['end_week']
                })
        else:
            # 没有课程名格式，尝试从整个块解析
            # 这种情况可能是格式不规范，尝试提取信息
            parsed = parse_fallback(block, "", weekday, default_start, default_end)
            if parsed['course_name']:
                courses.append(parsed)
    
    return courses


def parse_fallback(content: str, course_name: str, weekday: int, default_start: int, default_end: int) -> Dict[str, Any]:
    """备用解析方法"""
    # 周次
    week_match = re.search(r'(\d+)-(\d+)周(?:\((单|双)\))?', content)
    start_week, end_week, week_type = 1, 16, 'all'
    if week_match:
        start_week = int(week_match.group(1))
        end_week = int(week_match.group(2))
        if week_match.group(3) == '单':
            week_type = 'odd'
        elif week_match.group(3) == '双':
            week_type = 'even'
    
    # 节次
    section_match = re.search(r'(\d+)-(\d+)节', content)
    start_section = int(section_match.group(1)) if section_match else default_start
    end_section = int(section_match.group(2)) if section_match else default_end
    
    # 教师（教务系统格式）
    teacher_match = re.search(r'([^\[\s]+)\[(\d+),([^\]]+)\]', content)
    teacher = teacher_match.group(1).strip() if teacher_match else ""
    
    # 地点
    location = ""
    if teacher_match:
        after = content[teacher_match.end():].strip()
        loc_match = re.match(r'(一教\d+|科研楼\d+|[^\s]+训练场|[^\s]+场|\S+楼\d+)', after)
        if loc_match:
            location = loc_match.group(1)
    
    # 如果没有课程名，尝试从内容开头提取
    if not course_name:
        lines = content.strip().split('\n')
        if lines:
            first_line = lines[0].strip()
            # 尝试匹配课程名格式
            name_match = re.match(r'^([^\[\n]+)', first_line)
            if name_match:
                course_name = name_match.group(1).strip()
                # 清理可能的数字后缀
                course_name = re.sub(r'\s+\d+$', '', course_name)
    
    return {
        'course_name': course_name,
        'teacher': teacher,
        'location': location,
        'weekday': weekday,
        'start_section': start_section,
        'end_section': end_section,
        'week_type': week_type,
        'start_week': start_week,
        'end_week': end_week
    }


def parse_excel(file_path: str) -> Dict[str, Any]:
    """解析Excel课表"""
    result = {'success': False, 'message': '', 'courses': [], 'errors': []}
    
    try:
        df = build_dataframe_from_excel(file_path)
        weekday_mapping = detect_weekday_columns(df)
        section_mapping = build_row_section_mapping(df)
        all_courses = []
        
        for row_idx in range(len(df)):
            if row_idx not in section_mapping:
                continue
            start_section, end_section = section_mapping[row_idx]
            
            for col_idx in range(len(df.columns)):
                if col_idx not in weekday_mapping:
                    continue
                
                weekday = weekday_mapping[col_idx]
                cell = df.iloc[row_idx, col_idx]
                courses = parse_course_cell(cell, weekday, start_section, end_section)
                all_courses.extend(courses)
        
        # 去重
        seen = set()
        unique_courses = []
        for c in all_courses:
            key = (c['course_name'], c['weekday'], c['start_section'], c['end_section'],
                   c['week_type'], c['start_week'], c['end_week'], c['location'])
            if key not in seen:
                seen.add(key)
                unique_courses.append(c)
        
        unique_courses.sort(key=lambda x: (x['weekday'], x['start_section'], x['start_week']))
        
        result['courses'] = unique_courses
        result['success'] = True
        result['message'] = f'成功解析 {len(unique_courses)} 门课程'
        
    except Exception as e:
        result['errors'].append(f'文件读取失败: {str(e)}')
        result['message'] = 'Excel文件读取失败'
    
    return result


def update_config(courses: list) -> bool:
    """更新config.json"""
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
        config['courses'] = courses
        with open('config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"错误: 更新配置失败: {e}")
        return False


def main():
    if len(sys.argv) < 2:
        print("使用方法: python import_courses.py <课表Excel文件>")
        print("\n支持的格式：")
        print("  格式1：课程名[编号]")
        print("         周次 教师地点")
        print("  示例：高等数学[001]")
        print("        1-16周 示例老师A 示例教室A")
        sys.exit(1)
    
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在: {file_path}")
        sys.exit(1)
    
    print(f"正在解析: {file_path}\n")
    
    result = parse_excel(file_path)
    
    if result['errors']:
        print("解析错误:")
        for e in result['errors']:
            print(f"  - {e}")
    
    if not result['success']:
        print(f"\n解析失败: {result['message']}")
        sys.exit(1)
    
    print(f"{result['message']}\n")
    
    # 显示结果
    weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    for weekday in range(1, 8):
        day_courses = [c for c in result['courses'] if c['weekday'] == weekday]
        if day_courses:
            print(f"【{weekday_names[weekday - 1]}】")
            for c in day_courses:
                week_info = f"{c['start_week']}-{c['end_week']}周"
                if c['week_type'] == 'odd':
                    week_info += "(单)"
                elif c['week_type'] == 'even':
                    week_info += "(双)"
                print(f"  第{c['start_section']}-{c['end_section']}节 | {c['course_name']} | {c['teacher']} | {c['location']} | {week_info}")
            print()
    
    if update_config(result['courses']):
        print(f"[OK] 已更新 config.json，共导入 {len(result['courses'])} 门课程")
    else:
        print("[ERROR] 更新配置失败")
        sys.exit(1)


if __name__ == "__main__":
    main()
